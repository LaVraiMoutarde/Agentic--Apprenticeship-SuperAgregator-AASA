"""
Excel Exporter — exporte les résultats de recherche et ranking en Excel formaté.

Colonnes TASK 5 :
    Rank | Final Score | Embedding Score | LLM Global | LLM Tech | LLM Profile
    Title | Company | Location | URL | Source | Contract
    Explanation | Strengths | Weaknesses | Risks | Description

Deux modes d'export :
    export_search_results(matches: list[SearchResult])     → Phase 5 (sans LLM)
    export_ranked_offers(ranked: list[RankedOffer])        → Phase 6 (avec LLM)

Usage :
    exporter = ExcelExporter("exports/")
    path = exporter.export_ranked_offers(ranked, "resultats-scored.xlsx")
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from src.search.retriever import SearchResult
from src.scoring.llm_scorer import RankedOffer
from src.store.models import Offer


class ExcelExporter:
    """Exporte les résultats de recherche en Excel formaté."""

    def __init__(self, output_dir: str = "exports/") -> None:
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    # ── Export générique (liste d'Offer ORM) ──

    def export_offers(
        self,
        offers: list[Offer],
        filename: str = "offres.xlsx",
    ) -> Path:
        """Exporte une liste d'offres ORM en Excel."""
        if not offers:
            raise ValueError("Aucune offre à exporter")

        rows = []
        for o in offers:
            rows.append({
                "Title": o.title or "",
                "Company": o.company or "",
                "Location": o.location or "",
                "Region": o.region or "",
                "Contract": o.contract_type or "",
                "Domain": o.domain or "",
                "Level": o.required_level or "",
                "Source": o.source or "",
                "URL": o.url or "",
                "Score": o.data_quality_score or 0,
                "Scraped": o.scraped_date or "",
                "Description": (o.description or "")[:300],
            })

        df = pd.DataFrame(rows)
        path = self.output_dir / filename

        with pd.ExcelWriter(str(path), engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Offres", index=False)
            ws = writer.sheets["Offres"]
            widths = {
                "Title": 45, "Company": 25, "Location": 25, "Region": 18,
                "Contract": 15, "Domain": 18, "Level": 10, "Source": 14,
                "URL": 35, "Score": 8, "Scraped": 16, "Description": 40,
            }
            self._apply_widths(ws, df, widths)

        return path.resolve()

    # ── Phase 5 : export simple (sans LLM) ──

    def export_matches(
        self,
        matches: list[SearchResult],
        filename: str = "resultats.xlsx",
    ) -> Path:
        """Exporte les résultats de recherche sémantique (sans scoring LLM)."""
        if not matches:
            raise ValueError("Aucun résultat à exporter")

        rows = []
        for r in matches:
            o = r.offer
            rows.append({
                "Rank": r.rank,
                "Similarity": r.similarity_score,
                "Title": o.title or "",
                "Company": o.company or "",
                "Location": o.location or "",
                "URL": o.url or "",
                "Source": o.source or "",
                "Quality": o.data_quality_score or 0,
                "Contract": o.contract_type or "",
                "Description": (o.description or "")[:200],
            })

        df = pd.DataFrame(rows).sort_values("Similarity", ascending=False)
        path = self.output_dir / filename

        with pd.ExcelWriter(str(path), engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Matches", index=False)
            ws = writer.sheets["Matches"]
            widths = {
                "Rank": 8, "Similarity": 12, "Title": 45, "Company": 25,
                "Location": 25, "URL": 35, "Source": 15, "Quality": 10,
                "Contract": 15, "Description": 40,
            }
            self._apply_widths(ws, df, widths)

        return path.resolve()

    # ── Phase 6 : export avec scoring LLM (TASK 5) ──

    def export_ranked_offers(
        self,
        ranked: list[RankedOffer],
        filename: str = "resultats-scored.xlsx",
    ) -> Path:
        """Exporte les offres rankées avec scores LLM (TASK 5 output)."""
        if not ranked:
            raise ValueError("Aucun résultat à exporter")

        rows = [r.to_dict() for r in ranked]

        # Aplatir les listes pour Excel
        for row in rows:
            row["strengths"] = " | ".join(row["strengths"]) if row.get("strengths") else ""
            row["weaknesses"] = " | ".join(row["weaknesses"]) if row.get("weaknesses") else ""
            row["risks"] = " | ".join(
                f"[{r['type']}] {r['detail']}" for r in row.get("risks", [])
            ) if row.get("risks") else ""

        df = pd.DataFrame(rows)

        # Colonnes dans l'ordre TASK 5
        column_order = [
            "rank", "final_score", "embedding_score",
            "llm_global_score", "llm_technical_score", "llm_profile_score",
            "title", "company", "location", "url", "source", "contract_type",
            "explanation", "strengths", "weaknesses", "risks", "description_snippet",
        ]
        df = df[[c for c in column_order if c in df.columns]]

        path = self.output_dir / filename

        with pd.ExcelWriter(str(path), engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Ranked", index=False)
            ws = writer.sheets["Ranked"]
            widths = {
                "rank": 7, "final_score": 13, "embedding_score": 16,
                "llm_global_score": 16, "llm_technical_score": 17, "llm_profile_score": 17,
                "title": 45, "company": 25, "location": 25, "url": 35,
                "source": 14, "contract_type": 16,
                "explanation": 55, "strengths": 40, "weaknesses": 40,
                "risks": 45, "description_snippet": 45,
            }
            self._apply_widths(ws, df, widths)

        return path.resolve()

    # ── Helpers ──

    @staticmethod
    def _apply_widths(ws, df: pd.DataFrame, widths: dict[str, int]) -> None:
        """Applique les largeurs de colonnes et les filtres auto."""
        from openpyxl.utils import get_column_letter
        for i, col in enumerate(df.columns, 1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = widths.get(col, 15)
        last_col_letter = get_column_letter(len(df.columns))
        ws.auto_filter.ref = f"A1:{last_col_letter}{len(df) + 1}"
